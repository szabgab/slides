import openpyxl
import datetime

wb = openpyxl.Workbook()

ws = wb.active

ws['A1'] = 19
ws['A2'] = 23

ws['A3'] = "=A1+A2"

wb.save("expression.xlsx")

