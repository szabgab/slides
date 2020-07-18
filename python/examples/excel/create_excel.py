import openpyxl
import datetime

wb = openpyxl.Workbook()

ws = wb.active

ws['A1'] = 42

ws['A2'] = datetime.datetime.now()
#ws.column_dimensions['A'].width = 20.0

wb.save("first.xlsx")
