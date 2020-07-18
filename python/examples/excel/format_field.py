import openpyxl
import datetime

wb = openpyxl.Workbook()

ws = wb.active

ws['A1'] = 123456.78
ws['A2'] = 123456.78
ws['A3'] = 123456.78
ws['A4'] = -123456.78
ws['A5'] = datetime.datetime.now()
ws.column_dimensions['A'].width = 20.0

ws['A2'].number_format = '0.00E+00'
ws['A3'].number_format = '#,##0_);[RED](#,##0)'
ws['A4'].number_format = '#,##0_);[RED](#,##0)'

wb.save("format.xlsx")

