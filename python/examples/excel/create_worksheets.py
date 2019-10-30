import openpyxl
import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 42
ws.title = "First"

ws2 = wb.create_sheet()
ws2.title = "Second sheet"
ws2['A1'] = datetime.datetime.now()
ws2.sheet_properties.tabColor = "1072BA"

wb.save("two_worksheets.xlsx")

