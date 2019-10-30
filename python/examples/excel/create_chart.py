import openpyxl

wb = openpyxl.Workbook()

ws = wb.active
ws.title = "Chart"

a = ["First", 20, 28, 30, 37, 18, 47]
b = ["Second", 35, 30, 40, 40, 38, 35]

# write them as columns
for i in range(len(a)):
    ws.cell(row=i+1, column=1).value = a[i]
    ws.cell(row=i+1, column=2).value = b[i]

lc = openpyxl.chart.LineChart()
lc.title = "Two Lines Chart"
#lc.style=13
data = openpyxl.chart.Reference(ws,
                                min_col=1,
                                min_row=1,
                                max_col=2,
                                max_row=len(a))
lc.add_data(data, titles_from_data=True)

ws.add_chart(lc, "D1")
wb.save("chart.xlsx")
